import pandas as pd
import numpy as np 
import time
import os
import tkinter as tk
from collections import Counter
from tkinter import ttk, filedialog, messagebox
from tkinter import simpledialog
from datetime import datetime, timedelta
from datetime import datetime
import oss2
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import win32com
import win32com.client as win32
import re
from itertools import product
import gc
import subprocess
from configparser import ConfigParser
import threading
import sys





#定义程序参数集合

OSTRUS_DATE = '发情日期'
OSTRUS_TIME = '发情时间'
OSTRUS_FEMALE_ID = '发情母猪耳号'
OSTRUS_FEMALE_FRID = '发情母猪耳号-FRID'
OSTRUS_FEMALE_NAME = '发情母猪耳号文件'

MATING_UNIT = '配种单元'
MATING_LINE = '配种线别'
MATING_DATE = '配种日期'
MATING_TIME = '配种时间'
MATING_BATCH = '配种批次'



MB_EWB = '种猪选配二维表'
MB_SEMEN = '公猪采精信息汇总'
MB_SEMEN_DEMAN = '【模板】精液需求表'
MB_SELECTION_MATING = '【模板】猪只选配表'
MD_BATCH_RE = '批次关系.ini'
CONFIG_FILE_PATH = 'config.ini'


ID_LOCATION = "配怀种猪耳号信息"
FEMALE_ID = '母猪耳号'
MALE_ID = '公猪耳号'
VARIETY = '品种'
FEMALE_VARIETY = '母猪品种'
MALE_VARIETY = '公猪品种'
BOAR_NUMBER = '编号'
FIELD_LOCATION = '栏位'
SEMEN_GRANT_COPIES = '精液发放份数'
PLAN_RETAINED_SEMEN_COPIES =  '计划自留份数'
REMAIN_RETAINED_SEMEN_COPIES = '剩余自留份数'
USAGE_SEMEN_COPIES = '使用份数'
REMAIN_SEMEN_COPIES = '剩余份数'


COMPLETE_IND_NUM = '系统耳牌号'
QUANTITY = '数量'
UPLOAD_BASIC_INF = '基础信息上传'
SELECTION_MATING_PLAN = '选配方案'
DOWNLOAD_FILE = '文件下载'
MATING_LOCATION='贵州爱科欣 1 线'
COLOMNS_ID_CODE = 'Unnamed: 1'

MATCH_MATING_NUMBER = '可配母猪头数'
COEFFICIENT_RE = '亲缘系数'

RANK = '排名'
RANK_FEMALE = '母猪排名'
RANK_MALE = '公猪排名'
RANK_MALE_PERFORMANCES = '公猪性能排名'
RANK_MATCH_MATING_NUMBER = '可配头数排名'
RANK_COEFFICIENT_RE = '亲缘系数排名'

SELECTION_MATING_OPTIONAL = '选配情况'

PZ_START_DATE = '配种开始日期'
PZ_END_DATE = '配种结束日期'


BREED_CODE_MAP = {
    "YY": "大白",
    "LL": "长白",
    "DD": "杜洛克",
    "PP": "皮特兰",
    "PD": "皮杜"
}

LOCATION_CODE_MAP = {
    "一单元": "1",
    "二单元": "2",
    "三单元": "3",
    "四单元": "4",
    "五单元": "5"
}


#筛选亲缘系数3.125以下的选配情况
max_kinship_threshold = 3.125

def match_load(file_name):
    with open(file_name, 'r',encoding='utf-8') as file:
        content = file.read()
    lines = content.split('\n')
    columns = lines[0].split('\t')
    data = [line.split('\t') for line in lines[1:]]
    df = pd.DataFrame(data, columns=columns)
    return df

def config_load(file_path):
    # 创建解析器对象
    config = ConfigParser()
    config.read(file_path, encoding='utf-8')
    config_dict = {}
    for section in config.sections():
        # 将该 section 下的所有参数存储为字典
        section_dict = dict(config.items(section))
        # 将 section 字典存储在 config_dict 中
        config_dict[section] = section_dict
    return config_dict



#获取英文品种对应中文ID
def get_breed_code(variety):
    return BREED_CODE_MAP.get(variety, "")


#获取单元位置对应ID
def get_location_code(location):
    return LOCATION_CODE_MAP.get(location, "")



#上传文件至系统（原始档案及公母猪信息，由育种专员上传；精液分配和选配表，由系统自动上传）
def upload_file_to_oss(upload_file,file_path):
    if file_path:
        oss_path = f"{upload_file}{file_path.split('/')[-1]}"  # 服务器上的路径
        bucket.put_object_from_file(oss_path, file_path)
        if upload_file=="Mating_data/":
            messagebox.showinfo("提示", "选配文件上传成功")
    else:
        messagebox.showinfo("提示", "文件路径错误")
        return None

#获取数据存储所在的所有文件夹（原始数据）
def oss_file_all(prefix = 'data/'):
    full_path_list = []
    file_name_list = []
    for obj in oss2.ObjectIterator(bucket, prefix=prefix):
        if obj.key != prefix:  # 检查键是否等于 prefix，如果不等于则添加到列表中
            full_path_list.append(obj.key)
            file_name_list.append(os.path.basename(obj.key))
    return {'full_path':full_path_list,'file_name':file_name_list}


def oss_file_complete_xlsx(prefix = 'complete_data/选配配种记录表/xlsx/'):
    full_path_list = []
    file_name_list = []
    for obj in oss2.ObjectIterator(bucket, prefix=prefix):
        if obj.key != prefix:  # 检查键是否等于 prefix，如果不等于则添加到列表中
            full_path_list.append(obj.key)
            file_name_list.append(os.path.basename(obj.key))
    return {'full_path':full_path_list,'file_name':file_name_list}



#比较日期大小
def parse_date(match):
    if match is None:
        return datetime.min  # 返回一个特殊值，表示匹配失败
    date_str = match.group(0)
    return datetime.strptime(date_str, "%Y.%m.%d")

#比较日期并从服务器下载数据
def matching_files_change(matching_files):
    match_date = []
    for file_name in matching_files:
        match_date.append(re.match(r'(\d{4}\.\d{1,2}\.\d{1,2})', file_name))
    if pd.isna(match_date).all():
        bucket.get_object_to_file(f"{os.path.dirname(oss_file_all()['full_path'][0])}/{matching_files[0]}", f"{matching_files[0]}")
    else:
        latest_match = max(match_date, key=parse_date, default=None)                    #根据文件前缀盘点，日期最大值
        prefix_file_name = latest_match.group(0)                                        #获取日期最大值具体字符
        paths_with = [file for file in matching_files if prefix_file_name in file]        #筛选日期字符对应的文件
        contains_am = any('上午' in file for file in paths_with)                          #判断是否有上午数据
        contains_pm = any('下午' in file for file in paths_with)                          #判断是否有下午数据
        if contains_am and contains_pm:#若上午下午文件同时存在，则返回下午文件路径
            paths_with = [file for file in paths_with if '下午' in file]
        elif contains_am:
            pass
        file_name_date = f"{os.path.dirname(oss_file_all()['full_path'][0])}/{paths_with[0]}"#获取对应文件的下载路径
        bucket.get_object_to_file(file_name_date, f"{os.path.basename(file_name_date)}")     #从服务器下载对应数据


#选择合适的文件进行打开
def find_file(file_key_words):
        matching_files = glob.glob(f"*{file_key_words}*")
        if not matching_files:
            return None
        match_date = []
        for file_name in matching_files:
                match_date.append(re.match(r'(\d{4}\.\d{1,2}\.\d{1,2})', file_name))
        if pd.isna(match_date).all():
                file_name = f"{matching_files[0]}"
        else:
                latest_match = max(match_date, key=parse_date, default=None)                    #根据文件前缀盘点，日期最大值
                prefix_file_name = latest_match.group(0)                                        #获取日期最大值具体字符
                paths_with = [file for file in matching_files if prefix_file_name in file]        #筛选日期字符对应的文件
                contains_am = any('上午' in file for file in paths_with)                          #判断是否有上午数据
                contains_pm = any('下午' in file for file in paths_with)                          #判断是否有下午数据
                if contains_am and contains_pm:#若上午下午文件同时存在，则返回下午文件路径
                        paths_with = [file for file in paths_with if '下午' in file]
                elif contains_am:
                        pass
                file_name = f"{paths_with[0]}"
        return file_name


def get_parent_directory(file_path):
    return os.path.dirname(file_path)


#创建空Event事件，确保函数on_select可正常运行
class Event:
    pass

def on_select(event):
    file_name_to_full_path = dict(zip(oss_file_all()['file_name'], oss_file_all()['full_path']))
    selected_file_name = variety_combobox.get()
    selected_full_path = file_name_to_full_path[selected_file_name]
    return selected_full_path,selected_file_name

def down_file_from_oss_specify():
    selected_full_path = on_select(Event())[0]
    selected_file_name = on_select(Event())[1]
    bucket.get_object_to_file(f"{selected_full_path}", f"{selected_file_name}")


#自动下载当前云文档的数据
def download_file_from_oss():
    match_date = []
    matching_files = []
    matching_files1 = []
    for oss_path in oss_file_all()['full_path']:
        file_name = os.path.basename(oss_path)
        if MB_EWB not in file_name and MB_SEMEN not in file_name:
            bucket.get_object_to_file(oss_path, file_name)
        elif MB_EWB in file_name:
            matching_files = [os.path.basename(file) for file in oss_file_all()['full_path'] if f"{MB_EWB}({stored_data[VARIETY]})" in os.path.basename(file)]
        elif MB_SEMEN in file_name:
            matching_files1 = [os.path.basename(file) for file in oss_file_all()['full_path'] if f"{stored_data[MATING_BATCH]}{MB_SEMEN}" in os.path.basename(file)]
    if matching_files:
        matching_files_change(matching_files)
    if matching_files1:
        matching_files_change(matching_files1)


#检查耳号是否存在
def check_ear_numbers(ear_numbers, df,column_name):
    df_set = set(df[column_name].values)
    not_in_table =[ear for ear in ear_numbers if ear not in df_set]
    return not_in_table


#定义函数显示竖式提醒
def show_vertical_message(not_in_table,massage_al):
    message = f"{massage_al}，请更正：\n" + "\n".join(not_in_table)
    messagebox.showinfo("提示", message)


def mating_number_judge():
    mating_judege = [stored_data[OSTRUS_FEMALE_FRID],stored_data[OSTRUS_FEMALE_ID]]
    ear_numbers = list(filter(None,[judge for judge in mating_judege if len(judge) !=0][0].split('\n')))
    return ear_numbers

#数据检查
def data_checking(stored_data):
    
   #采取异步下载文件的方式，避免程序卡顿
    download_thread = threading.Thread(target=download_file_from_oss)
    download_thread.start()	#开始下载所需文件
    download_thread.join()  # 确保下载完成再继续执行

    #发情母猪耳号输入判断
    female_name = stored_data.get(OSTRUS_FEMALE_NAME, "")
    female_id = stored_data.get(OSTRUS_FEMALE_ID, "")
    if female_name and not os.path.isfile(female_name):
        messagebox.showerror("路径错误", "文件名无效，请重新输入文件名或选择有效文件。")
        return None
    if female_name and female_id:
        messagebox.showinfo("提示", "请选择或输入正确的发情母猪耳号，系统不支持两者同时输入")
        return None
    if not female_name and not female_id:
        messagebox.showinfo("提示", "请选择发情母猪FRID文件 或 输入发情母猪完整15位个体号")
        return None
    if os.path.exists(female_name):
        mating_date_file = os.path.basename(female_name)
        oss_files = oss_file_all(prefix='Mating_data/')['file_name']
        if mating_date_file not in oss_files:
            messagebox.showerror("提示", "请点击<上传>按钮，上传选配耳号文件。")
            return None
    ear_numbers = mating_number_judge()


    #参数输入判断
    keys = [key for key in stored_data.keys() if key != OSTRUS_FEMALE_ID and key != OSTRUS_FEMALE_FRID and key != OSTRUS_FEMALE_NAME]
    for key in keys:
        if len(stored_data[key])==0:
            messagebox.showinfo("提示", f"请输入{key}")
            return None

   
    #判断品种是否选择正确（对应文件是否存在）
    file_two_dimensional_table = find_file(f"{stored_data[MATING_BATCH]}{MB_EWB}({stored_data[VARIETY]})")
    if file_two_dimensional_table:
        df_two_dimensional_table = pd.read_excel(file_two_dimensional_table)
    else:
        messagebox.showinfo("提示", '检测到无对应选配二维表，请选择正确的品种和批次')
        return None

    #获取基本批次配种参数
    file_name_pzjl = f"{stored_data[MATING_BATCH]}【模板】配种记录.txt"
    mating_limit =config_dict['database']['mating_limit']

    #若单词输入耳号大于配种批次设定数，提醒用户重新输入
    if len(ear_numbers) > float(mating_limit):
        messagebox.showinfo("提示", "本次配种数量超过批次设定最大值（60），请再次确认配种耳号是否正确")
        return None

    #判断批次配种量，若大于80%，则每头公猪预留精液为0，否则每头公猪预留精液量为
    global condition
    condition = 2
    if os.path.exists(file_name_pzjl):
        df = open_and_read_txt(file_name_pzjl)
        if len(ear_numbers)+len(df)>=float(mating_limit)*0.8:
            condition = 0
        if len(ear_numbers)+len(df)-1>float(mating_limit):
            if not messagebox.askokcancel("提示", f"本次配种数{len(ear_numbers)}\n\n加上本次配种，该批次配种数{len(ear_numbers)+len(df)}\n\n已超过60，是否继续进行选配？"):
                return None
            else:
                pass
            
        #判断当前耳号是否已选配
        in_table = [i for i in ear_numbers if i in df[FEMALE_ID].values]
        if in_table:
            show_vertical_message(in_table,'该个体号已存在配种选配记录')
            return None

    #判断配种个体是否有选配信息(在二维表中匹配配种耳号)
    not_in_table = check_ear_numbers(ear_numbers, df_two_dimensional_table,COLOMNS_ID_CODE)
    if not_in_table:
        show_vertical_message(not_in_table,'未找到个体号选配信息')
        return None

    #判断生产线是否正确（匹配文件是否存在）
    file_Fields = find_file(f"{stored_data[MATING_LINE]}{ID_LOCATION}")
    if not file_Fields:
        messagebox.showinfo("提示", '请选择正确的生产线')
        return None
    
    #判断栏位是否正确（在单元内存栏匹配配种耳号）
    df_Fields  = pd.read_excel(file_Fields, sheet_name = stored_data[MATING_UNIT])
    # 找出不在DataFrame中的ear_numbers
    not_in_table = check_ear_numbers(ear_numbers, df_Fields,COMPLETE_IND_NUM)
    if not_in_table:
        show_vertical_message(not_in_table,'未找到个体号栏位信息，请确认配种单元或耳号是否正确')
        return None
    
    #判断所选日期当日是否存在配种选配记录（与云平台文件进行对比）
    global excel_file_name_selective_breeding
    global excel_file_name_Semen_distribution
    oss_file_name =oss_file_complete_xlsx()['file_name']
    now_file_name = f"1.2一线{stored_data[OSTRUS_DATE][5:]}{stored_data[OSTRUS_TIME]}发情，{stored_data[MATING_DATE]}{stored_data[MATING_TIME]}（初配+复配）猪只选配表.xlsx"
    base_name = f"1.2一线{stored_data[OSTRUS_DATE][5:]}{stored_data[OSTRUS_TIME]}发情，{stored_data[MATING_DATE]}{stored_data[MATING_TIME]}（初配+复配）"
    suffix = ""
    if now_file_name in oss_file_name:
        if not messagebox.askokcancel("提示", "当前日期已存在对应配种记录，是否继续进行选配操作？"):
            messagebox.showinfo("提示", '请重新核实并选择配种时间')
            return None
        if not messagebox.askokcancel("提示", "是否覆盖原配种选配记录？"):
            suffix = "1"
    excel_file_name_selective_breeding = f"{base_name}猪只选配表{suffix}.xlsx"
    excel_file_name_Semen_distribution = f"{base_name}精液需求表{suffix}.xlsx"
    return True

def close_and_remove_files():
    gc.collect()
    del_file_name = [file for file in oss_file_all()['file_name'] if os.path.exists(file)]
    for file in del_file_name:
        subprocess.run(["del", file], check=True, shell=True)



#定义并保存数据，全局变量
def save_data():

    del_thread = threading.Thread(target=close_and_remove_files)
    del_thread.start()
    del_thread.join()


    global stored_data
    stored_data = submit_data()


    #耳号判断
    file = stored_data[OSTRUS_FEMALE_NAME]
    stored_data[OSTRUS_FEMALE_FRID] = ''
    if os.path.exists(file):
        try:
            df_mating_data = pd.read_excel(file,sheet_name ='Sheet1')
            stored_data[OSTRUS_FEMALE_FRID] = '\n'.join(map(str, df_mating_data[COMPLETE_IND_NUM]))
        except Exception:
            messagebox.showerror("错误", "请选择正确的发情母猪耳号文件")
            return None
    
    if data_checking(stored_data):
        messagebox.showinfo("提示", '个体号验证成功')
        return stored_data
    else:
        return None

# 保存用户输入数据的函数
def submit_data():

    data = {OSTRUS_FEMALE_ID: data_entry_input.get("1.0", tk.END).strip(),
            OSTRUS_FEMALE_NAME: data_entry_select.get(),
            MATING_UNIT: unit_combobox.get(),
            VARIETY:veriaty_combobox.get(),
            OSTRUS_DATE:oestrus_date_combobox.get(),
            OSTRUS_TIME:oestrus_time_combobox.get(),
            MATING_DATE:mating_date_combobox.get(),
            MATING_TIME:mating_time_combobox.get(),
            MATING_LINE:Breeding_Line_combobox.get(),
            MATING_BATCH:Breeding_batch_combobox.get(),}
    return data


#选择文件按钮对应参数
def choose_file(ind_entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path and os.path.isfile(file_path):
        ind_entry.delete(0, tk.END)
        ind_entry.insert(0, file_path)
    else:
        messagebox.showerror("路径错误", "选留个体信息文件路径无效，请重新选择有效文件。")
        return None


comboboxes = []
#根据选择情况，更新下拉选择框
def update_comboboxes(event):
    current_selection = Module_combobox.get()
    clear_frame(dynamic_frame)
    if UPLOAD_BASIC_INF == current_selection:
        Window_addition_File_selection()
    elif SELECTION_MATING_PLAN == current_selection:
        Window_addition_basic_selective_breeding()
    else:
        Window_addition_all_selection()

#清楚表单
def clear_frame(frame):
    for widget in frame.winfo_children():
        widget.destroy()

#
def Window_addition_all_selection():
    row_Fields = '5'
    Module_name = '文件名'
    Window_addition_basic_all_selection(row_Fields, Module_name)

#基础信息上传模块，模块生成
def Window_addition_File_selection():
    global date_combobox
    selective_breeding_Fields = 3
    selective_breeding = '二维选配表'
    Boar_semen_Fields = 4
    Boar_semen = "公猪精液可用情况"
    Nonpregnant_Fields = 5
    Nonpregnant_name = "待配母猪栏位信息："
    Window_addition_basic_File_selection(Nonpregnant_Fields, Nonpregnant_name)
    Window_addition_basic_File_selection(Boar_semen_Fields, Boar_semen)
    Window_addition_basic_File_selection(selective_breeding_Fields, selective_breeding)
    #添加上传日期下拉选择框
    row_date = 2
    formatted_date = datetime.now().strftime("%Y.%m.%d")
    current_date = datetime.strptime(formatted_date, "%Y.%m.%d")
    date_range = [current_date - timedelta(days=i) for i in range(7, -8, -1)]
    formatted_dates = [date.strftime("%Y.%m.%d") for date in date_range]
    tk.Label(dynamic_frame, text=OSTRUS_DATE).grid(row=row_date, column=0, sticky=tk.W, padx=5, pady=5)
    date_combobox = ttk.Combobox(dynamic_frame, values=formatted_dates, width=26)
    date_combobox.grid(row=row_date, column=1, sticky=tk.W, padx=5, pady=5)
    date_combobox.set(formatted_date)  # 设置默认值

#基础信息上传模块，模块生成基础
def Window_addition_basic_File_selection(row_Fields, Module_name):
    ttk.Label(dynamic_frame, text=Module_name).grid(row=row_Fields, column=0, sticky=tk.W, padx=5, pady=5)
    ind_entry = ttk.Entry(dynamic_frame, width=29)
    ind_entry.grid(row=row_Fields, column=1, sticky=tk.W, padx=5, pady=5, columnspan=1)
    choose_file_button = ttk.Button(dynamic_frame, text="选择文件",
                                    command=lambda entry=ind_entry: choose_file(entry))
    choose_file_button.grid(row=row_Fields, column=2, padx=5, pady=5)
    upload_button = ttk.Button(dynamic_frame, text="上传",
                               command=lambda file_path=ind_entry.get(): upload_file_to_oss('data/',ind_entry.get()))
    upload_button.grid(row=row_Fields, column=3, padx=5, pady=5)
    return ind_entry

#文件下载模块
def Window_addition_basic_all_selection(row_Fields, Module_name):
    global variety_combobox
    ttk.Label(dynamic_frame, text=Module_name).grid(row=row_Fields, column=0, sticky=tk.W, padx=5, pady=5)
    variety_combobox = ttk.Combobox(dynamic_frame, values=oss_file_all()['file_name'], width=30)
    variety_combobox.grid(row=row_Fields, column=1, sticky=tk.W, padx=5, pady=5)
    variety_combobox.set("")  # 设置默认值
    # 添加下载按钮
    confirm_button = ttk.Button(dynamic_frame, text="下载", command=down_file_from_oss_specify)
    confirm_button.grid(row=row_Fields, column=2, padx=5, pady=5, columnspan=1)



def update_breeding_batch(event=None):
    pzxb = config_dict['database']['pzxb'].split(',')
    xb_code = config_dict['database']['xb_code'].split(',')
    pzxb_dict = dict(zip(pzxb, xb_code))
    MD_BATCH_RE = os.path.join(base_path, '批次关系.ini')
    df_Module_date = match_load(MD_BATCH_RE)
    df_Module_date[PZ_START_DATE] = pd.to_datetime(df_Module_date[PZ_START_DATE])
    df_Module_date[PZ_END_DATE] = pd.to_datetime(df_Module_date[PZ_END_DATE])
    time_late = timedelta(days=6)
    formatted_date = pd.to_datetime('today')  # 示例日期，可根据实际情况调整
    index = df_Module_date[
        (df_Module_date[PZ_START_DATE] - time_late <= formatted_date) &
        (df_Module_date[PZ_END_DATE] + time_late >= formatted_date)
        ]
    Module_date = list(index[BOAR_NUMBER])

    if event:
        selected_pzxb = Breeding_Line_combobox.get()
        pzxb_code = pzxb_dict[selected_pzxb]
        filtered_data = [ind for ind in Module_date if pzxb_code in ind]
        Breeding_batch_combobox['values'] = filtered_data
        Breeding_batch_combobox.set(filtered_data[-1])  # 重置批次选择框的值

    return pzxb, Module_date
#选配方案模块
def Window_addition_basic_selective_breeding():
    global data_entry_input
    global data_entry_select
    global unit_combobox
    global oestrus_date_combobox
    global oestrus_time_combobox
    global mating_date_combobox
    global mating_time_combobox
    global veriaty_combobox
    global Breeding_Line_combobox
    global Breeding_batch_combobox
    #日期格式转换
    formatted_date = datetime.now().strftime("%Y/%m/%d")
    current_date = datetime.strptime(formatted_date, "%Y/%m/%d")
    formatted_date_default = str(current_date.year) + '年' + str(current_date.month) + '月' + str(current_date.day) + '日' #默认数据
    formatted_time_default = str(current_date.month) + '月' + str(current_date.day) + '日' #默认数据
    date_range = [current_date - timedelta(days=i) for i in range(15, -15, -1)]
    formatted_dates_data = [str(date.year) + '年' + str(date.month) + '月' + str(date.day) + '日' for date in date_range]
    formatted_dates_time = [str(date.month) + '月' + str(date.day) + '日' for date in date_range]

    #配种单元选择框
    row_module = 1
    Module_unit = ['一单元','二单元','三单元','四单元','五单元']
    tk.Label(dynamic_frame, text=MATING_UNIT).grid(row=row_module, column=0, sticky=tk.W, padx=5, pady=5)
    unit_combobox = ttk.Combobox(dynamic_frame, values=Module_unit, width=26)
    unit_combobox.grid(row=row_module, column=1, sticky=tk.W, padx=5, pady=5)
    unit_combobox.set('一单元')  # 设置默认值

    #配种单元选择框
    row_module = 1
    Module_unit = ['大白','长白','杜洛克','皮特兰']
    tk.Label(dynamic_frame, text=VARIETY).grid(row=row_module, column=2, sticky=tk.W, padx=5, pady=5)
    veriaty_combobox = ttk.Combobox(dynamic_frame, values=Module_unit, width=25)
    veriaty_combobox.grid(row=row_module, column=3, sticky=tk.W, padx=5, pady=5)
    veriaty_combobox.set('')  # 设置默认值

    #添加上传日期下拉选择框
    row_date = 2
    tk.Label(dynamic_frame, text=OSTRUS_DATE).grid(row=row_date, column=0, sticky=tk.W, padx=5, pady=5)
    oestrus_date_combobox = ttk.Combobox(dynamic_frame, values=formatted_dates_data, width=26)
    oestrus_date_combobox.grid(row=row_date, column=1, sticky=tk.W, padx=5, pady=5)
    oestrus_date_combobox.set(formatted_date_default)  # 设置默认值

    #添加配种日期输入框
    row_module = 2
    Module_date = ['上午','下午']
    tk.Label(dynamic_frame, text=OSTRUS_TIME).grid(row=row_module, column=2, sticky=tk.W, padx=5, pady=5)
    oestrus_time_combobox = ttk.Combobox(dynamic_frame, values=Module_date, width=25)
    oestrus_time_combobox.grid(row=row_module, column=3, sticky=tk.W, padx=5, pady=5)
    oestrus_time_combobox.set('')  # 设置默认值

    row_date = 3

    tk.Label(dynamic_frame, text=MATING_DATE).grid(row=row_date, column=0, sticky=tk.W, padx=5, pady=5)
    mating_date_combobox = ttk.Combobox(dynamic_frame, values=formatted_dates_time, width=26)
    mating_date_combobox.grid(row=row_date, column=1, sticky=tk.W, padx=5, pady=5)
    mating_date_combobox.set(formatted_time_default)  # 设置默认值

    #添加配种日期输入框
    row_module = 3
    Module_date = ['上午','下午']
    tk.Label(dynamic_frame, text=MATING_TIME).grid(row=row_module, column=2, sticky=tk.W, padx=5, pady=5)
    mating_time_combobox = ttk.Combobox(dynamic_frame, values=Module_date, width=25)
    mating_time_combobox.grid(row=row_module, column=3, sticky=tk.W, padx=5, pady=5)
    mating_time_combobox.set('')  # 设置默认值
    
    #添加配种批次输入框
    row_module = 4
    pzxb, Module_date = update_breeding_batch()
    
    tk.Label(dynamic_frame, text=MATING_LINE).grid(row=row_module, column=0, sticky=tk.W, padx=5, pady=5)
    Breeding_Line_combobox = ttk.Combobox(dynamic_frame, values=pzxb, width=26)
    Breeding_Line_combobox.grid(row=row_module, column=1, sticky=tk.W, padx=5, pady=5)
    Breeding_Line_combobox.set('')  # 设置默认值
    Breeding_Line_combobox.bind("<<ComboboxSelected>>", update_breeding_batch)

    #添加配种批次输入框
    row_module = 4
    tk.Label(dynamic_frame, text=MATING_BATCH).grid(row=row_module, column=2, sticky=tk.W, padx=5, pady=5)
    Breeding_batch_combobox = ttk.Combobox(dynamic_frame, values=Module_date, width=25)
    Breeding_batch_combobox.grid(row=row_module, column=3, sticky=tk.W, padx=5, pady=5)
    Breeding_batch_combobox.set('')  # 设置默认值

    # 添加输入框
    row_date = 6
    ttk.Label(dynamic_frame, text=OSTRUS_FEMALE_ID).grid(row=row_date, column=0, padx=5, pady=5)
    # 添加一个多行输入框
    data_entry_input = tk.Text(dynamic_frame, height=10, width=25)
    data_entry_input.grid(row=row_date, column=1, padx=5, pady=5)
    # 添加确认按钮
    confirm_button = ttk.Button(dynamic_frame, text="确认", command=save_data)
    confirm_button.grid(row=row_date, column=2, padx=5, pady=5)
    # 添加确认按钮
    confirm_button = ttk.Button(dynamic_frame, text="生成选配文件", command=selective_breeding)
    confirm_button.grid(row=row_date, column=3, padx=5, pady=5)

#######发情母猪耳号文件选择框#######
    # 添加输入框
    row_date = 5
    ttk.Label(dynamic_frame, text=OSTRUS_FEMALE_ID).grid(row=row_date, column=0, padx=5, pady=5)#添加标签
    data_entry_select = ttk.Entry(dynamic_frame, width=29)#添加输入框
    data_entry_select.grid(row=row_date, column=1, sticky=tk.W, padx=5, pady=5, columnspan=1)
    #选择文件按钮
    choose_file_button = ttk.Button(dynamic_frame, text="选择文件",command=lambda entry=data_entry_select: choose_file(entry))
    choose_file_button.grid(row=row_date, column=2, padx=5, pady=5)
    
    upload_button = ttk.Button(dynamic_frame, text="上传",
                               command=lambda file_path=data_entry_select.get(): upload_file_to_oss('Mating_data/',data_entry_select.get()))
    upload_button.grid(row=row_date, column=3, padx=5, pady=5)



#根据输入的个体信息，在二维选配表中构建选配情况
def Estrus_selection():
    """
    选择发情母猪与种公猪进行配对并计算精液使用量。
    根据实际二维表进行评估选择，个别没有参配的母猪单独标记

    返回:
        dict: 包含以下键的字典：
            - 'positions1': 参配种公猪耳号列表
            - 'positions2': 参配种母猪耳号
            - 'positions_counter_doubled': 每个公猪的精液计划使用量（数量乘以2）
            - 'Unselected_individual': 未做选配的发情母猪列表
    """
            
            
    ear_numbers = mating_number_judge()
    
    df_two_dimensional_table  = open_two_dimensional_table()

    matched_rows  = df_two_dimensional_table[df_two_dimensional_table[COLOMNS_ID_CODE].isin(ear_numbers)]

    positions1 = []
    positions2 = []
    positions_location = []
    Unselected_individual = []
    #前置函数已判断输入耳号全部在二维表中显示，则对公母猪选配信息进行位置查找和汇总
    for index, row in matched_rows.iterrows():
        for column, value in row.items():
            if isinstance(value, str) and value.startswith('###') and value.endswith('###'):
                positions1.append(df_two_dimensional_table .loc[(0,column)])
                positions2.append(df_two_dimensional_table .loc[(index, COLOMNS_ID_CODE)])
                positions_location.append((index,column))
        if all(not isinstance(item, str) or not item.startswith('###') for item in list(row)):
            Unselected_individual.append(list(row)[1])
    #公猪信息汇总
    positions_counter = Counter(positions1)
    #精液计划使用量计算
    positions_counter_doubled = {key: value * 2 for key, value in positions_counter.items()}
    return {'positions1':positions1,'positions2':positions2,'positions_counter_doubled':positions_counter_doubled,'Unselected_individual':Unselected_individual}
def semen_count(semen_enough):
    
    counts = semen_enough[MALE_ID].value_counts() * 2
    
    #获取计数列
    df_Semen_usage = counts.reset_index()
    
    #列命名
    df_Semen_usage.columns = [MALE_ID, SEMEN_GRANT_COPIES]
    df_Semen_usage = df_Semen_usage.reset_index(drop=True)
    
    return df_Semen_usage
#查找并返回指定坐标
# 自定义的find_positions函数
def find_positions(df, not_selected_F, semen_M):
    positions = []

    for i in range(df.shape[0]):
        if df.iloc[i, 1] in not_selected_F:
            for j in range(df.shape[1]):
                if df.iloc[0, j] in semen_M:
                    positions.append((i, j))

    return positions
#从工作表查找数据
def find_position(sheet, value):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value == value:
                return (row, col)
    return None
def selection_process(df_combinations_process,mating_max_count):

    #设置为每头公猪最多额外配种母猪
    boar_remaining = {boar:min(mating_max_count,remaining) for boar,
                      remaining in df_combinations_process[[MALE_ID, MATCH_MATING_NUMBER]]
                      .drop_duplicates().set_index(MALE_ID)[MATCH_MATING_NUMBER].items()}
    #设置空列表
    selected_sows = set()
    #根据选配规则进行选配
    for idx, row in df_combinations_process.iterrows():
        sow_ear = row[FEMALE_ID]
        boar_ear = row[MALE_ID]
        # 如果母猪还没有被配种且公猪有剩余可配头数
        if sow_ear not in selected_sows and boar_remaining[boar_ear] > 0:
            df_combinations_process.at[idx, SELECTION_MATING_OPTIONAL] = '√'
            selected_sows.add(sow_ear)
            boar_remaining[boar_ear] -= 1

    return selected_sows,df_combinations_process
#统计目录下某列有颜色的单元格数量
def count_color(file,col):
    workbook = load_workbook(file)
    worksheet = workbook.active
    colored_cell_count = 0
    for cell in worksheet[col]:
        if cell.fill.start_color.index != '00000000':  # 如果单元格有填充颜色
            colored_cell_count += 1
    return colored_cell_count
#中文日期格式转换
def chinese_date_change(date):
    match = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date)

    if match:
        year = match.group(1)
        month = match.group(2)
        day = match.group(3)
        # 格式化为"YYYY.MM.DD"
        formatted_date = f"{year}.{month}.{day}"
    return formatted_date

def xlsx_change_pdf(excel_file_name, pdf_file_name):
    try:
        # 获取当前脚本文件所在文件夹的路径
        script_dir = os.path.dirname(os.path.abspath(__file__))

        # 构建完整的 Excel 文件路径和 PDF 文件路径
        excel_file_path = os.path.join(script_dir, excel_file_name)
        pdf_file_path = os.path.join(script_dir, pdf_file_name)
        # 检查 Excel 文件是否存在
        if not os.path.exists(excel_file_path):
            print(f"文件路径错误：{excel_file_path} 不存在")
            return

        # 启动 Excel 应用程序
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False  # 可选，设置为 False 以隐藏 Excel 窗口

        # 打开工作簿
        workbook = excel.Workbooks.Open(excel_file_path)

        # 选择要保存为 PDF 的工作表（假设是第一个工作表）
        #worksheet = workbook.Worksheets['Sheet1']  # 假设要转换的是第一个工作表
        # 保存为 PDF
        workbook.ExportAsFixedFormat(Type=0, Filename=f"{pdf_file_path}.pdf", Quality=0)

    except Exception as e:
        print(f"发生错误：{e}")

    finally:
        # 确保工作簿和 Excel 应用程序关闭
        try:
            workbook.Close(SaveChanges=False)
            excel.Quit()
        except:
            pass




def open_Boar_semen():
    #读取文件
    file_paths_Boar_semen = find_file(f"【模板】{stored_data[MATING_BATCH]}{MB_SEMEN}")
    df_Boar_semen = pd.read_excel(file_paths_Boar_semen)
    return df_Boar_semen
def open_two_dimensional_table():
    file_two_dimensional_table = find_file(f"{MB_EWB}({stored_data[VARIETY]})")
    df_two_dimensional_table = pd.read_excel(file_two_dimensional_table)
    return df_two_dimensional_table


def open_and_read_txt(file_name):
    with open(file_name, 'r') as file:
        content = file.read()
    lines = content.split('\n')
    columns = lines[0].split('\t')
    data = [line.split('\t') for line in lines[1:]]
    df = pd.DataFrame(data, columns=columns)
    return df

def assign_selection_status(group):
    # 获取当前组的可配母猪头数并将其转换为整数
    num_assignments = int(group.iloc[0][MATCH_MATING_NUMBER])  # 可配母猪头数

    # 按照‘排名’列对组内数据进行升序排序
    group = group.sort_values(by=RANK)

    # 初始化‘选配情况’列为空字符串
    group[SELECTION_MATING_OPTIONAL] = ''

    # 将前num_assignments行的‘选配情况’列值设置为‘√’
    group.iloc[:num_assignments, group.columns.get_loc(SELECTION_MATING_OPTIONAL)] = '√'

    # 返回更新后的组
    return group

def save_dataframe_to_txt(df, filename):
    """
    逐行将 DataFrame 数据追加到 TXT 文件中，避免多余换行。

    参数:
    - df: 要写入的 DataFrame
    - filename: 目标文件名
    """
    with open(filename, 'a') as f:
        for index, row in df.iterrows():
            line = '\t'.join(row.astype(str)) + '\n'
            if f.tell() == 0:  # 如果文件为空，写入表头
                header = '\t'.join(df.columns) + '\n'
                f.write(header)
            f.write(line)


#选配情况判断
class SemenSupplyManager:
    def __init__(self):
        self.Cache_Semen_supply = False
        self.df_name_enough = None
        self.df_name_Insufficient = None
        self.df_Semen_usage = None
    def Semen_supply(self):
        """
	根据精液供应情况和母猪耳号进行种猪选配，返回选配结果和精液发放情况。

    	返回:
    	
    	self.df_name_enough (DataFrame): 母猪选配情况，结构为['公猪耳号','母猪耳号']
    	
    	self.df_name_Insufficient(DataFrame)：精液不足的选配情况，结构为['公猪耳号','母猪耳号']
    	
    	df_Semen_usage(DataFrame)，公猪精液使用情况，结构为['公猪耳号','精液发放份数']

	"""
        if not self.Cache_Semen_supply:
            
            self.Cache_Semen_supply = True
            # 获取公猪精液需求汇总数据
            positions_counter_doubled = Estrus_selection()['positions_counter_doubled']

            # 生成初步精液需求表
            self.df_Semen_usage = pd.DataFrame(list(positions_counter_doubled.items()), columns=[MALE_ID, SEMEN_GRANT_COPIES])
            # 打开公猪精液需求表（最后保存）
            df_Boar_semen = open_Boar_semen()
            # 将精液需求和精液情况进行合并
            df_merge = pd.merge(self.df_Semen_usage, df_Boar_semen, on=MALE_ID, how='outer')

            # 若初次使用，即剩余份数为空，则填充剩余自留份数
            if pd.isna(df_merge[REMAIN_SEMEN_COPIES]).all():
                df_merge[REMAIN_SEMEN_COPIES] = df_merge[REMAIN_RETAINED_SEMEN_COPIES]
            else:
                df_merge[REMAIN_RETAINED_SEMEN_COPIES] = df_merge[REMAIN_SEMEN_COPIES]

            df_merge.fillna(0, inplace=True)
            
            #对本批次选配情况进行判断若本批次配种数已达90%，则剩余母猪使用排名考前的公猪进行配种
            if condition == 0:
                not_selected_table = pd.DataFrame(Estrus_selection()['positions2'],columns = [FEMALE_ID])
                if len(not_selected_table) <= 5:
                    df_two_dimensional_table = open_two_dimensional_table()
                    #索引第0行，从1开始的所有公猪耳号
                    df_male = df_two_dimensional_table.loc[0, 1:].to_dict()
                    df_female = df_two_dimensional_table.loc[df_two_dimensional_table[COLOMNS_ID_CODE].isin(not_selected_table[FEMALE_ID]), COLOMNS_ID_CODE].to_dict()
                    coordinate_p = pd.DataFrame(product(df_female,df_male),columns = [RANK_FEMALE,RANK_MALE])
                    values_p = pd.DataFrame(product(df_female.values(),df_male.values()),columns = [FEMALE_ID,MALE_ID])
                    #数据表格进行拼接
                    df_selecting_mating = pd.concat([values_p,coordinate_p], axis=1)
                    #获取对应亲缘系数并转换为浮点数
                    df_selecting_mating[COEFFICIENT_RE] = [float(item.replace('###', '').rstrip('%'))
                                                           for item in
                                                           [df_two_dimensional_table.at[female, male]
                                                            for female, male in zip(df_selecting_mating[RANK_FEMALE], df_selecting_mating[RANK_MALE])]]

                    df_selecting_mating = pd.merge(df_selecting_mating, df_Boar_semen[[MALE_ID, REMAIN_SEMEN_COPIES]], on=MALE_ID, how='outer')
                    
                    df_selecting_mating = df_selecting_mating[df_selecting_mating[REMAIN_SEMEN_COPIES]>2]

                    df_selecting_mating[MATCH_MATING_NUMBER] = np.trunc((df_selecting_mating[REMAIN_SEMEN_COPIES]-condition) / 2)

                    df_selecting_mating[RANK_COEFFICIENT_RE] = df_selecting_mating.groupby(FEMALE_ID)[COEFFICIENT_RE].rank(method='min')

                    df_selecting_mating[RANK_MATCH_MATING_NUMBER] = df_selecting_mating.groupby(FEMALE_ID)[MATCH_MATING_NUMBER].rank(method='min', ascending=False)

                    df_selecting_mating[RANK_MALE_PERFORMANCES] = df_selecting_mating.groupby(FEMALE_ID)[RANK_MALE].rank(method='min')

                    df_selecting_mating[RANK] = df_selecting_mating[RANK_MALE_PERFORMANCES]*0.5 + df_selecting_mating[RANK_COEFFICIENT_RE]*0.3 + df_selecting_mating[RANK_MATCH_MATING_NUMBER]*0.2

                    df_selecting_mating = df_selecting_mating.sort_values(by=[RANK_FEMALE,RANK]).reset_index(drop=True)

                    df_selecting_mating[SELECTION_MATING_OPTIONAL] = ''
                    #筛选亲缘系数3.125以下的选配情况
                    max_kinship_threshold = 3.125

                    df_selecting_mating_process = df_selecting_mating[df_selecting_mating[COEFFICIENT_RE] <= max_kinship_threshold]
                    if df_selecting_mating_process.empty:
                        if messagebox.askokcancel("提示", f"当前（近交系数：3.125）下，无精液可用，是否重新设定近交系数"):
                            root = tk.Tk()
                            root.withdraw()  # 隐藏主窗口
                        while True:
                            # 弹出输入框
                            new_threshold = simpledialog.askfloat("设定新阈值", "请输入新的近交系数阈值:")
                            if new_threshold is not None:
                                max_kinship_threshold = new_threshold
                                df_selecting_mating_process = df_selecting_mating[df_selecting_mating[COEFFICIENT_RE] <= max_kinship_threshold]
                                break
                            else:
                                messagebox.showwarning("提示", "输入无效，请输入一个有效的数值。")
                    selected_sows,df_selecting_mating_process = selection_process(df_selecting_mating_process,10)

                    mark1 = df_selecting_mating_process[SELECTION_MATING_OPTIONAL] == '√'
                    
                    self.df_name_enough = df_selecting_mating_process[mark1][[MALE_ID, FEMALE_ID]].reset_index(drop=True)
                    
                    self.df_name_Insufficient = not_selected_table[~not_selected_table[FEMALE_ID].isin(self.df_name_enough[FEMALE_ID])].reset_index(drop=True)

                    self.df_Semen_usage = semen_count(self.df_name_enough)
                    
                    return self.df_name_enough, self.df_name_Insufficient, self.df_Semen_usage,2

            # 将公母猪耳号合并
            df_position = pd.DataFrame(zip(Estrus_selection()['positions1'], Estrus_selection()['positions2']), columns=[MALE_ID, FEMALE_ID])


            name_Insufficient = df_merge[df_merge[SEMEN_GRANT_COPIES] > df_merge[REMAIN_SEMEN_COPIES]-2][MALE_ID]
            # 生成精液不足的配种情况
            self.df_name_Insufficient = df_position[df_position[MALE_ID].isin(name_Insufficient)]

            # 精液足够的配种情况
            self.df_name_enough = df_position[~df_position[MALE_ID].isin(name_Insufficient)]

            # 重置索引
            self.df_name_enough = self.df_name_enough.reset_index(drop=True)
            self.df_name_Insufficient = self.df_name_Insufficient.reset_index(drop=True)

            # 获取公猪精液使用份数统计
            self.df_Semen_usage = semen_count(self.df_name_enough)

            # 返回参数，精液足够的选配情况，精液不足的个体选配情况，精液发放情况
            return self.df_name_enough, self.df_name_Insufficient, self.df_Semen_usage,1
        else:
            # 返回缓存的结果
            return self.df_name_enough, self.df_name_Insufficient, self.df_Semen_usage,1




def Semen_supply_Insufficient_Preprocessing(df_name_Insufficient):
    """
    根据精液供应情况和母猪耳号进行种猪选配预处理，返回选配结果、精液发放情况、未选配情况。

    返回:
    selected_table (DataFrame): 母猪选配情况，结构为['公猪耳号','母猪耳号']
    Semen_usage(DataFrame),公猪精液使用情况，结构为['公猪耳号','精液发放份数']
    not_selected_table(DataFrame):未选配的母猪个体号
    """
    if df_name_Insufficient.empty:
        return None
    else:
        df_Boar_semen = open_Boar_semen()
        #默认留2份精液
        df_Boar_semen = df_Boar_semen[df_Boar_semen[REMAIN_SEMEN_COPIES]>2]
        if df_Boar_semen.empty:
            return None,None,df_name_Insufficient,False
        number_Boar_semen = pd.merge(df_name_Insufficient, df_Boar_semen, on=MALE_ID, how='inner')[[MALE_ID, FEMALE_ID, REMAIN_SEMEN_COPIES]]
        
  
        number_Boar_semen[MATCH_MATING_NUMBER] = np.trunc((number_Boar_semen[REMAIN_SEMEN_COPIES]-2)/2)

        #对二维表进行操作 		  #模糊匹配获取二维表路径
        df_two_dimensional_table = open_two_dimensional_table()
        #获取选配母猪耳号的排名
        indexes = df_two_dimensional_table.index[df_two_dimensional_table[COLOMNS_ID_CODE]
                                                 .isin(number_Boar_semen[FEMALE_ID])].tolist()
        #合并母猪耳号和排名信息
        F_Index = pd.DataFrame(zip(number_Boar_semen[FEMALE_ID], indexes), columns=[FEMALE_ID, RANK])
        #将母猪信息和公猪信息合并
        number_Boar_semen_merge = pd.merge(number_Boar_semen, F_Index, on=FEMALE_ID, how='outer')
        #选择排名靠前的个体进行配种
        number_Boar_semen_merge = number_Boar_semen_merge.groupby(MALE_ID).apply(assign_selection_status)
        #重置索引顺序
        number_Boar_semen_merge = number_Boar_semen_merge.reset_index(drop=True)
        #已选配个体
        selected_table = number_Boar_semen_merge[number_Boar_semen_merge[SELECTION_MATING_OPTIONAL] == '√'][[MALE_ID, FEMALE_ID]]

        #未选配个体
        not_selected_table = df_name_Insufficient[~df_name_Insufficient[FEMALE_ID].isin(selected_table[FEMALE_ID])]

        Semen_usage = semen_count(selected_table)

        return selected_table, Semen_usage, not_selected_table,False

def semen_supply_process(df_male,df_combinations,df_Boar_semen,df_two_dimensional_table):
    global max_kinship_threshold
    global semen_count_fre
    df_female = df_two_dimensional_table[df_two_dimensional_table[COLOMNS_ID_CODE].isin(df_combinations[FEMALE_ID])].iloc[:, 1].to_dict()
    coordinate_p = pd.DataFrame(product(df_female,df_male),columns = [RANK_FEMALE,RANK_MALE])
    
    values_p = pd.DataFrame(product(df_female.values(),df_male.values()),columns = [FEMALE_ID,MALE_ID])
    
    df_selecting_mating = pd.concat([values_p,coordinate_p], axis=1)
    #获取对应亲缘系数并转换为浮点数
    df_selecting_mating[COEFFICIENT_RE] = [float(item.replace('###', '').rstrip('%'))
                                           for item in
                                           [df_two_dimensional_table.at[female, male]
                                            for female, male in zip(df_selecting_mating[RANK_FEMALE], df_selecting_mating[RANK_MALE])]]


    df_selecting_mating = pd.merge(df_selecting_mating, df_Boar_semen, left_on=MALE_ID, right_on=MALE_ID, how='inner')

    
    #正常情况下的判断
    
    if semen_count_fre <=10:
        df_selecting_mating = df_selecting_mating[df_selecting_mating[REMAIN_SEMEN_COPIES]>2]
        df_selecting_mating[MATCH_MATING_NUMBER] = np.trunc((df_selecting_mating[REMAIN_SEMEN_COPIES]-2)/2)
    else:
        df_selecting_mating = df_selecting_mating[df_selecting_mating[REMAIN_SEMEN_COPIES]>0]
        df_selecting_mating[MATCH_MATING_NUMBER] = np.trunc((df_selecting_mating[REMAIN_SEMEN_COPIES])/2)
    semen_count_fre+=1
    df_selecting_mating[RANK_COEFFICIENT_RE] = df_selecting_mating.groupby(FEMALE_ID)[COEFFICIENT_RE].rank(method='min')

    df_selecting_mating[RANK_MATCH_MATING_NUMBER] = df_selecting_mating.groupby(FEMALE_ID)[MATCH_MATING_NUMBER].rank(method='min', ascending=False)

    df_selecting_mating[RANK] = df_selecting_mating[RANK_COEFFICIENT_RE]*0.6 + df_selecting_mating[RANK_MATCH_MATING_NUMBER]*0.4

    df_selecting_mating = df_selecting_mating.sort_values(by=[RANK_FEMALE,RANK]).reset_index(drop=True)

    df_selecting_mating[SELECTION_MATING_OPTIONAL] = ''


    df_selecting_mating_process = df_selecting_mating[df_selecting_mating[COEFFICIENT_RE] <= max_kinship_threshold]

    if df_selecting_mating_process.empty:
        if messagebox.askokcancel("提示", f"当前（近交系数：3.125）下，无精液可用，是否重新设定近交系数"):
            root = tk.Tk()
            root.withdraw()  # 隐藏主窗口
            #循环
            while True:
                # 弹出输入框
                new_threshold = simpledialog.askfloat("设定新阈值", "请输入新的近交系数阈值:")
                if new_threshold is not None:
                    max_kinship_threshold = new_threshold
                    df_selecting_mating_process = df_selecting_mating[df_selecting_mating[COEFFICIENT_RE] <= max_kinship_threshold]
                    break
                else:
                    messagebox.showwarning("提示", "输入无效，请输入一个有效的数值。")
        else:
            return None,None,True
    
    #设置每头公猪最大配种头数为2
    selected_sows,df_selecting_mating_process = selection_process(df_selecting_mating_process,2)
    return selected_sows,df_selecting_mating_process,False

#对精液不足情况下进行判断
semen_count_fre = 0
def Semen_supply_Insufficient(not_selected_table):
    """
    根据精液供应情况和母猪耳号进行种猪选配，返回选配结果和精液发放情况。

    返回:
    semen_enough (DataFrame): 母猪选配情况，结构为['公猪耳号','母猪耳号']
    df_Semen_usage(DataFrame),公猪精液使用情况，结构为['公猪耳号','精液发放份数']
    not_selected_table(DataFrame):未选配的母猪个体号
    """
    global semen_count_fre
    df_Boar_semen = open_Boar_semen()[[MALE_ID, REMAIN_SEMEN_COPIES]]
    
    if (df_Boar_semen[REMAIN_SEMEN_COPIES]<=3).all().all():
        messagebox.showinfo("提示", "当前无精液可用，请更新精液信息")
        file_name = f"{stored_data[OSTRUS_DATE]}{stored_data[OSTRUS_TIME]}{stored_data[MATING_BATCH]}发情未配耳号.txt"
        not_selected_table = pd.DataFrame(df_name_Insufficient[FEMALE_ID])
        not_selected_table['未配种原因'] = '精液不足'
        save_dataframe_to_txt(not_selected_table, file_name)
        return None,None,not_selected_table,True
    df_two_dimensional_table = open_two_dimensional_table()

    
    #筛选份数大于2公猪

    filtered_df = df_Boar_semen[df_Boar_semen[REMAIN_SEMEN_COPIES] > 0]
    
    #若无精液可以使用，则跳出循环
    if filtered_df.empty:
        messagebox.showinfo("提示", "当前无精液（精液库存低）可供使用，请更新可用精液信息")
        return None,None,not_selected_table,True
    #筛选排名前五的公猪精液剩余量个体
    if semen_count_fre <=5:
        semen_M = filtered_df.nlargest(5, REMAIN_SEMEN_COPIES)[[MALE_ID, REMAIN_SEMEN_COPIES]]
    elif semen_count_fre >5 and semen_count_fre <=10:
        semen_M = filtered_df.nlargest(min(10,len(filtered_df)), REMAIN_SEMEN_COPIES)[[MALE_ID, REMAIN_SEMEN_COPIES]]
    else:
        semen_M = filtered_df.nlargest(len(filtered_df), REMAIN_SEMEN_COPIES)[[MALE_ID, REMAIN_SEMEN_COPIES]]
    semen_count_fre+=1
    #获取未做选配的母猪个体
    not_selected_F = not_selected_table[FEMALE_ID]

    #将所有待配母猪和精液公猪组合
    combinations = list(product(not_selected_F, semen_M[MALE_ID]))

    df_combinations = pd.DataFrame(combinations, columns=[FEMALE_ID, MALE_ID])

    #获取参配公猪序号及对应个体号
    df_male = df_two_dimensional_table.iloc[0, :][df_two_dimensional_table.iloc[0, :].isin(df_combinations[MALE_ID])].to_dict()

    selected_sows,df_selecting_mating_process,jypd = semen_supply_process(df_male,df_combinations,df_Boar_semen,df_two_dimensional_table)
    if jypd:
        return None,None,not_selected_table,True
    #根据选配情况更新not_selected_table
    semen_enough = df_selecting_mating_process[df_selecting_mating_process[SELECTION_MATING_OPTIONAL] == '√'][[MALE_ID, FEMALE_ID]]

    #已选配个体
    selected_mother_ear_numbers = semen_enough[FEMALE_ID].unique()

    #未选配个体
    not_selected_table =not_selected_table[~not_selected_table[FEMALE_ID].isin(semen_enough[FEMALE_ID])]

    #对公猪精液使用情况进行统计
    df_Semen_usage = semen_count(semen_enough)
    jlkcpd = False
    return semen_enough,df_Semen_usage,not_selected_table,jlkcpd



#针对未做精准选配个体，使用后5头公猪进行配种
def Semen_supply_Unselected_individual(Unselected_individual):
    """
    对未做选配的个体进行精准选配。

    返回:
        tuple: 包含以下数据框的列表：
            - result_Unselected(DataFrame): 选配情况，结构为['公猪耳号','母猪耳号']
            - df_Semen_usage(DataFrame): 精液发放情况，结构为['公猪耳号','精液发放份数']
            - unable_Unselected(numpy)：未作选配的个体号，结构为['母猪耳号']
    """
    
    df_Boar_semen = open_Boar_semen()[[MALE_ID, REMAIN_SEMEN_COPIES]]
    #构建df
    Unselected_ID = pd.DataFrame({FEMALE_ID:Unselected_individual})
    #读取二维表
    df_two_dimensional_table = open_two_dimensional_table()
    #读取列索引值（后5头公猪）
    df_male = df_two_dimensional_table.iloc[0,-5:].to_dict()
    #获取行索引值（母猪）
    df_Unselected_individual = pd.DataFrame(Unselected_individual,columns = [FEMALE_ID])
    
    selected_sows,df_selecting_mating_process,jypd = semen_supply_process(df_male,df_Unselected_individual,df_Boar_semen,df_two_dimensional_table)
    if jypd:
        return None,None,Unselected_ID,True
    result_Unselected = df_selecting_mating_process[df_selecting_mating_process[SELECTION_MATING_OPTIONAL]=='√'][[MALE_ID, FEMALE_ID]].reset_index(drop=True)

    selected_mother_ear_numbers = result_Unselected[FEMALE_ID].unique()

    not_selected_table = df_Unselected_individual [~df_Unselected_individual [FEMALE_ID].isin(selected_mother_ear_numbers)]

    
    df_Semen_usage = semen_count(result_Unselected)



    #返回参数，选配情况
    return result_Unselected,df_Semen_usage,not_selected_table



#每次对精液使用情况进行汇总，并对精液进行调整
def semen_usage(df_Semen_usage):
    if pd.isna(df_Semen_usage).all().all():
        return None
    file_paths_Boar_semen = find_file(f"{stored_data[MATING_BATCH]}{MB_SEMEN}") #通过字段匹配，获取完整路径名
    df_Boar_semen = open_Boar_semen()
    #将公猪精液使用和精液信息汇总情况合并
    df_merge = pd.merge(df_Boar_semen, df_Semen_usage, on=MALE_ID, how='outer')
    df_semen_distribution= df_merge[[BOAR_NUMBER, MALE_ID, SEMEN_GRANT_COPIES]]
    df_semen_distribution= df_semen_distribution.sort_values(by=BOAR_NUMBER)
    df_merge = df_merge.sort_values(by=BOAR_NUMBER).reset_index(drop=True)
    if pd.isna(df_merge[REMAIN_SEMEN_COPIES]).all():
        pass
    else:
        df_merge[REMAIN_RETAINED_SEMEN_COPIES] = df_merge[REMAIN_SEMEN_COPIES]
    df_merge.fillna(0, inplace=True)
    df_merge[USAGE_SEMEN_COPIES] = df_merge[SEMEN_GRANT_COPIES]
    df_merge[REMAIN_SEMEN_COPIES] = df_merge[REMAIN_RETAINED_SEMEN_COPIES]-df_merge[SEMEN_GRANT_COPIES]

    workbook_Boar_semen = load_workbook(file_paths_Boar_semen)
    worksheet_Boar_semen = workbook_Boar_semen['Sheet1']
    worksheet_Boar_semen.delete_rows(3, worksheet_Boar_semen.max_row)

    start_row = 2  # 开始填入数据的起始行
    for index, row in df_merge.iterrows():
        for col, value in enumerate(row[[BOAR_NUMBER, MALE_ID, VARIETY, PLAN_RETAINED_SEMEN_COPIES, REMAIN_RETAINED_SEMEN_COPIES, USAGE_SEMEN_COPIES, REMAIN_SEMEN_COPIES]], 1):
            worksheet_Boar_semen.cell(row=start_row + index, column=col).value = value
            
    start_index = file_paths_Boar_semen.find('【')
    file_paths_Boar_semen = file_paths_Boar_semen[start_index:] if start_index != -1 else file_paths_Boar_semen 
    #模板重命名
    date_Boar_semen = stored_data[OSTRUS_DATE]
    
    file_Boar_semen = f"{chinese_date_change(date_Boar_semen)}{stored_data[OSTRUS_TIME]}{file_paths_Boar_semen}"
    workbook_Boar_semen.save(file_Boar_semen)
    upload_file_to_oss(f"data/",file_Boar_semen)

#读取公猪采精信息，并对公猪精液信息进行填入
def semen_distribution():
    """
	根据用户输入的母猪耳号进行选配操作，直至选配操作完毕。
	
	★★★执行流程★★★：
		1、利用SemenSupplyManager()对选配耳号进行操作，返回<选配结果（精液足够情况下）>、<精液不足的选配情况>、精液使用情况>
		2、针对精液不足的选配情况，利用Semen_supply_Insufficient_Preprocessing(df_name_Insufficient)进行操作，判断对应公猪可配头数，并选择排名靠前的母猪进行配种，返回<选配情况>、<未完成选配个体>、<精液使用情况>。
		3、针对未完成选配个体，利用Semen_supply_Insufficient(not_selected_table)进行选配操作，选择精液剩余量最大的5头公猪进行选配（每头公猪最大配种2头），利用亲缘系数、需求量等进行筛选，返回<选配情况>、<未完成选配个体>、<精液使用情况>
		4、重复过程3，直至不存在<未完成选配个体>
		5、针对未在选配表内进行选配标注是个体，利用Semen_supply_Unselected_individual(Unselected_individual)进行选配操作，选择排名靠后的5头公猪进行配种操作，返回<选配情况>、<未完成选配个体>、<精液使用情况>
		6、针对过程5产生的<未完成选配个体>，重复过程3，直至


		注：上述每个步骤生成的精液使用情况，都需要利用semen_usage(df_Semen_usage)对精液使用情况进行操作，并返回精液使用后的结果，方便后续进行精液使用判断。
    	返回:

    	result_selection(DataFrame): 母猪选配结果，结构为['公猪耳号','母猪耳号']

    	result_semen(DataFrame)：公猪精液使用情况，结构为['公猪耳号','精液发放份数']


    """
    try:
        #初始化存储器
        selected_table_Preprocessing = pd.DataFrame(columns=[MALE_ID, FEMALE_ID])
        unable_Unselected_process = pd.DataFrame(columns=[MALE_ID, FEMALE_ID])
        result_Unselected = pd.DataFrame(columns=[MALE_ID, FEMALE_ID])
        Semen_enough_In_process = pd.DataFrame(columns=[MALE_ID, FEMALE_ID])

        Semen_usage_In_process = pd.DataFrame(columns=[MALE_ID, SEMEN_GRANT_COPIES])
        unable_Semen_usage_process = pd.DataFrame(columns=[MALE_ID, SEMEN_GRANT_COPIES])


        #生成常规选配对应母猪选配结果、精液不足的选配情况、精液发放情况
        manager = SemenSupplyManager()

        #返回正常选配情况，或返回非正常选配及未选配耳号
        Selection_results_routine, Selection_Insufficient_routine, Semen_usage_routine,semen_pd =  manager.Semen_supply()
        if Selection_results_routine is None and Selection_Insufficient_routine is None and Semen_usage_routine is None:
            return result_Unselected,unable_Semen_usage_process,None,True
        else:
            semen_usage(Semen_usage_routine)

        if not Selection_Insufficient_routine.empty:
            #针对精液不足的选配情况，优先查看原计划公猪是否有精液可用（第一步判断时为总体判断，若单头公猪计划配种大于剩余量，也会放到此步做判断）
            #返回的结果为，选配情况、精液发放情况、未选配耳号情况
            if semen_pd == 2:
                selected_table_Preprocessing, Semen_usage_Preprocessing, not_selected_table_Preprocessing,jlkcpd = Semen_supply_Insufficient(Selection_Insufficient_routine)
                if jlkcpd:
                    return Selection_results_routine,Semen_usage_routine,not_selected_table_Preprocessing,True
                else:
                    semen_usage(Semen_usage_Preprocessing)
            elif semen_pd == 1:
                selected_table_Preprocessing, Semen_usage_Preprocessing, not_selected_table_Preprocessing,jlkcpd = Semen_supply_Insufficient_Preprocessing(Selection_Insufficient_routine)
                if jlkcpd:
                    return Selection_results_routine,Semen_usage_routine,not_selected_table_Preprocessing,True
                else:
                    semen_usage(Semen_usage_Preprocessing)

            #重复利用选配规则，对未选配耳号进行循环选配，直至选配完毕
            while not not_selected_table_Preprocessing.empty:
                    Semen_enough_Insufficient,Semen_usage_Insufficient,not_selected_table_Preprocessing,jlkcpd = Semen_supply_Insufficient(not_selected_table_Preprocessing)
                    if jlkcpd:
                        break####对精液不足的情况进行判断
                    Semen_enough_In_process = pd.concat([Semen_enough_In_process,Semen_enough_Insufficient], ignore_index=True)
                    semen_usage(Semen_usage_Insufficient)


        #针对未选配猪只，单独做选配程序
        Unselected_individual = Estrus_selection()['Unselected_individual']
        if Unselected_individual:
            result_Unselected,Semen_usage_Unselected,unable_Unselected = Semen_supply_Unselected_individual(Unselected_individual)
            semen_usage(Semen_usage_Unselected)
            # 针对 unable_Unselected 进行循环选配
            while not unable_Unselected.empty:
                Semen_enough_Unselected, Semen_usage_Unselected, unable_Unselected,jlkcpd = Semen_supply_Insufficient(unable_Unselected)
                if jlkcpd:
                    break####对精液不足的情况进行判断
                unable_Unselected_process = pd.concat([unable_Unselected_process, Semen_enough_Unselected], ignore_index=True)
                semen_usage(Semen_usage_Unselected)

        #生成的结果进行合并
        result_selection = result_selection = pd.concat([df for df in [Selection_results_routine,
                                                                       selected_table_Preprocessing,Semen_enough_In_process,
                                                                       result_Unselected,unable_Unselected_process] if not df.empty], ignore_index=True)

        result_semen = semen_count(result_selection)
        return result_selection,result_semen,None,False
    except Exception as e:
        messagebox.showerror("错误", f"发生了一个错误: {e}")
        return None,None,None,False


#选配方案,正常选配流程
def selective_breeding():
    """
        精液足够的公猪，优先发放精液，制作选配表和精液需求表，并在二维表中标记对应配种情况
        
        从semen_distribution()获取精液选配操作过程的结果：选配结果，精液使用结果，程序执行结果

        根据程序执行结果进行代码执行，若选配未正常完成，则将未完成耳号写入txt记事本

        操作步骤：
        	1、根据选配结果，生成选配记录，并写入批次配种记录txt
        	2、根据选配结果，生成选配表，并上传云端
        	3、根据精液使用结果，生成精液需求表，并上传云端
        	4、根据选配结果，对二维表进行标记，标记对应选配记录
        	5、将选配表和精液需求表转换为pdf，删除原excel表，并上传云端
        	6、提醒用户"程序执行完毕→→选配方案已生成"，并关闭程序

    """
    try:


        #加载猪只选配表
        file_paths_selective_breeding  = find_file(MB_SELECTION_MATING) # 查找选配表文件路径
        workbook_selective_breeding = load_workbook(file_paths_selective_breeding)# 加载选配表工作簿
        worksheet_selective_breeding = workbook_selective_breeding['Sheet1']# 选择工作表


        #加载精液需求表
        file_paths_Semen_distribution = find_file(MB_SEMEN_DEMAN) # 查找选配表文件路径
        workbook_Semen_distribution = load_workbook(file_paths_Semen_distribution)# 加载选配表工作簿
        worksheet_Semen_distribution = workbook_Semen_distribution['Sheet1']# 选择工作表

	#获取精液信息
        df_position,Semen_usage_routine,not_selected_table,xppd = semen_distribution() # 调用精液分配函数，获取选配及精液分配信息
        if df_position.empty:# 如果精液信息为空，则停止后续处理
            return None
        #若选配判断提前结束（精液不足），则执行下述代码
        if xppd:
            Unselected_individual = Estrus_selection()['Unselected_individual']# 获取未配种个体
            df_Unselected_individual = pd.DataFrame([not_selected_table,Unselected_individual],columns = [FEMALE_ID]) # 转换为DataFrame
            df_Unselected_individual['未配种原因'] = '精液不足'
            file_name_Unselected = f"{stored_data[OSTRUS_DATE]}{stored_data[OSTRUS_TIME]}{stored_data[MATING_BATCH]}发情未配耳号.txt"
            save_dataframe_to_txt(df_Unselected_individual, file_name_Unselected)
        
        # 生成配种记录表
        file_name_pzjl = f"{stored_data[MATING_BATCH]}【模板】配种记录.txt"  # 生成配种记录表文件名
        df_position_pzjl = df_position
        df_position[MATING_BATCH] = stored_data[MATING_BATCH]  # 添加配种批次
        df_position[MATING_DATE] = stored_data[MATING_DATE]  # 添加配种日期
        df_position[OSTRUS_DATE] = stored_data[OSTRUS_DATE]  # 添加发情日期
        df_position = df_position[[OSTRUS_DATE, MATING_DATE, MATING_BATCH, FEMALE_ID, MALE_ID]]  # 选择需要的列

         # 保存配种记录表
        save_dataframe_to_txt(df_position, file_name_pzjl)
        upload_file_to_oss(f"data/", file_name_pzjl)  # 上传到云盘

        # 打开精液需求表
        df_semen_distribution = open_Boar_semen()  # 调用函数打开精液需求表

        # 读取对应单元的耳号信息
        file_paths_selective_breeding = find_file(f"{stored_data[MATING_LINE]}{ID_LOCATION}")  # 查找耳号信息文件路径
        df_process_selective_breeding = pd.read_excel(file_paths_selective_breeding, sheet_name = stored_data[MATING_UNIT])  # 读取数据

        # 根据配种情况匹配栏位信息，拥有栏位、系统耳牌号信息
        df_process_selective_location = df_process_selective_breeding[df_process_selective_breeding[COMPLETE_IND_NUM].isin(df_position[FEMALE_ID])][[FIELD_LOCATION, COMPLETE_IND_NUM]]

        # 将配种信息和栏位信息进行合并
        df_selective_breeding = pd.merge(df_process_selective_location, df_position, left_on=COMPLETE_IND_NUM, right_on=FEMALE_ID, how='inner')
        df_selective_breeding = pd.merge(df_semen_distribution, df_selective_breeding, on=MALE_ID, how='inner')[[FIELD_LOCATION, FEMALE_ID, BOAR_NUMBER, MALE_ID]]
        
        # 将信息进行合并，获取公猪耳号、系统耳牌号、母猪耳号、栏位、编号
        df_selective_breeding[FEMALE_VARIETY] = df_selective_breeding[FEMALE_ID].str[:2]  # 获取母猪品种代码
        df_selective_breeding[FEMALE_VARIETY] = [get_breed_code(i) for i in df_selective_breeding[FEMALE_VARIETY]]
        df_selective_breeding[FEMALE_ID] = df_selective_breeding[FEMALE_ID].str[-8:]  # 获取母猪耳号
        df_selective_breeding[MALE_VARIETY] = df_selective_breeding[MALE_ID].str[:2]  # 获取公猪品种代码
        df_selective_breeding[MALE_VARIETY] = [get_breed_code(i) for i in df_selective_breeding[MALE_VARIETY]]
        df_selective_breeding[MALE_ID] = df_selective_breeding[MALE_ID].str[-8:]  # 获取公猪耳号
        df_selective_breeding[FIELD_LOCATION] = get_location_code(stored_data[MATING_UNIT]) + '-' + df_selective_breeding[FIELD_LOCATION]  # 添加栏位代码
        df_selective_breeding = df_selective_breeding.sort_values(by=FIELD_LOCATION).reset_index(drop=True)  # 排序和重置索引
   
        start_row = 4  # 开始填入数据的起始行
        #对猪只选配表进行填充
        for index, row in df_selective_breeding.iterrows():
            worksheet_selective_breeding.cell(row=start_row + index, column=2).value = row[FIELD_LOCATION]    #单元格内容填充
            worksheet_selective_breeding.cell(row=start_row + index, column=3).value = row[FEMALE_ID]
            worksheet_selective_breeding.cell(row=start_row + index, column=4).value = row[FEMALE_VARIETY]
            worksheet_selective_breeding.cell(row=start_row + index, column=5).value = row[BOAR_NUMBER]
            worksheet_selective_breeding.cell(row=start_row + index, column=6).value = row[MALE_ID]
            worksheet_selective_breeding.cell(row=start_row + index, column=7).value = row[MALE_VARIETY]
            worksheet_selective_breeding.cell(row=start_row + index, column=8).value = stored_data[MATING_DATE]
            worksheet_selective_breeding.cell(row=start_row + index, column=9).value = stored_data[MATING_TIME]

        #对A1进行数据填充
        worksheet_selective_breeding.cell(row=1, column=1).value = f"{MATING_LOCATION}{get_location_code(stored_data[MATING_UNIT])}单元 选配配种记录表"
        #对B1进行数据填充
        worksheet_selective_breeding.cell(row=2, column=1).value = f"{stored_data[OSTRUS_DATE]}{stored_data[OSTRUS_TIME]}发情，{stored_data[MATING_DATE]}{stored_data[MATING_TIME]}（初配+复配）"
        #文件保存为excel
        workbook_selective_breeding.save(excel_file_name_selective_breeding)

        #获取数据填充需要的数据汇总
        result_Semen_distribution = df_selective_breeding.groupby([BOAR_NUMBER, MALE_ID, MALE_VARIETY]).size().reset_index(name=QUANTITY)
        start_row = 4  # 开始填入数据的起始行
        #对精液需求表进行填充
        for index, row in result_Semen_distribution.iterrows():
            worksheet_Semen_distribution.cell(row=start_row + index, column=2).value = row[BOAR_NUMBER]        #单元格内容填充
            worksheet_Semen_distribution.cell(row=start_row + index, column=3).value = row[MALE_ID]
            worksheet_Semen_distribution.cell(row=start_row + index, column=4).value = row[MALE_VARIETY]
            worksheet_Semen_distribution.cell(row=start_row + index, column=5).value = row[QUANTITY]
            worksheet_Semen_distribution.cell(row=start_row + index, column=6).value = row[QUANTITY]
        worksheet_Semen_distribution.cell(row=2, column=1).value = f"{stored_data[OSTRUS_DATE]}{stored_data[OSTRUS_TIME]}发情，{stored_data[MATING_DATE]}{stored_data[MATING_TIME]}（初配+复配）"
        workbook_Semen_distribution.save(excel_file_name_Semen_distribution)

        #将精液需求表文件上传云盘
        upload_file_to_oss(f"complete_data/精液分配表/xlsx/",excel_file_name_Semen_distribution)

        #将选配表文件上传云盘
        upload_file_to_oss(f"complete_data/选配配种记录表/xlsx/",excel_file_name_selective_breeding)

        #对二维表进行操作
        file_two_dimensional_table = find_file(f"{MB_EWB}({stored_data[VARIETY]})")         #模糊匹配获取二维表路径
        workbook_two_dimensional_table = load_workbook(file_two_dimensional_table)           #加载二维表工作簿
        worksheet_two_dimensional_table = workbook_two_dimensional_table['Sheet1']              #加载二维表工作表
        # 查找交点位置并标记为黄色
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   #设定颜色代码，此处设置为绿色，RBG通道
        #通过定位公猪、母猪耳号获取到sheet内坐标，对坐标表格进行颜色填充
        for index, row in df_position.iterrows():
            boar_id = row[MALE_ID]
            sow_id = row[FEMALE_ID]
            boar_pos = find_position(worksheet_two_dimensional_table, boar_id)
            sow_pos = find_position(worksheet_two_dimensional_table, sow_id)
            if boar_pos and sow_pos:
                intersection = (sow_pos[0],boar_pos[1])  # 交点的行号来自母猪耳号，列号来自公猪耳号；索引为（行，列）
                cell = worksheet_two_dimensional_table.cell(row=intersection[0], column=intersection[1])
                cell.fill = green_fill  #对单元格进行颜色填充
        date_two_dimensional_table = stored_data[OSTRUS_DATE]
        start_index = file_two_dimensional_table.find('【')
        if start_index != -1:
            file_two_dimensional_table = file_two_dimensional_table[start_index:]
        file_name_workbook_two_dimensional_table = f"{chinese_date_change(date_two_dimensional_table)}{stored_data[OSTRUS_TIME]}{file_two_dimensional_table}"
        workbook_two_dimensional_table.save(file_name_workbook_two_dimensional_table)     #文件保存
        upload_file_to_oss(f"data/",file_name_workbook_two_dimensional_table)
        close_and_remove_files()
        #change_xlsx_to_pdf()
        messagebox.showinfo("提示", "选配方案已生成")
        root.destroy()
    except Exception as e:
        messagebox.showerror("错误", f"发生了一个错误: {e}")

def change_xlsx_to_pdf():
    #定义文件名
    pdf_file_name_selective_breeding = f"1.2一线{stored_data[OSTRUS_DATE][5:]}{stored_data[OSTRUS_TIME]}发情，{stored_data[MATING_DATE]}{stored_data[MATING_TIME]}（初配+复配）猪只选配表"
    #excel文件转换为pdf
    xlsx_change_pdf(excel_file_name_selective_breeding,pdf_file_name_selective_breeding)
    upload_file_to_oss(f"complete_data/选配配种记录表/pdf/",f"{pdf_file_name_selective_breeding}.pdf")

    #定义文件名
    pdf_file_name_Semen_distribution = f"1.1一线{stored_data[OSTRUS_DATE][5:]}{stored_data[OSTRUS_TIME]}发情，{stored_data[MATING_DATE]}{stored_data[MATING_TIME]}（初配+复配）精液需求表"
    #精液需求表转换
    xlsx_change_pdf(excel_file_name_Semen_distribution,pdf_file_name_Semen_distribution)
    upload_file_to_oss(f"complete_data/精液分配表/pdf/",f"{pdf_file_name_Semen_distribution}.pdf")
    messagebox.showinfo("提示", "选配表及精液需求表已生成并上传")
    
    #subprocess.run(["del", excel_file_name_Semen_distribution], check=True, shell=True)
    #subprocess.run(["del", excel_file_name_selective_breeding], check=True, shell=True)

def create_tk():
    global Module_combobox
    global dynamic_frame
    # 创建主窗口
    root = tk.Tk()
    root.title("选配计划")

    # 创建一个表单框架
    form_frame = ttk.Frame(root, padding="10 10 10 10")
    form_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # 创建模块选择下拉框
    row_module = 0
    Module_selection = [UPLOAD_BASIC_INF, SELECTION_MATING_PLAN, DOWNLOAD_FILE]
    tk.Label(form_frame, text="模块").grid(row=row_module, column=0, sticky=tk.W, padx=5, pady=5)
    Module_combobox = ttk.Combobox(form_frame, values=Module_selection, width=30)
    Module_combobox.grid(row=row_module, column=1, sticky=tk.W, padx=5, pady=5)
    Module_combobox.bind("<<ComboboxSelected>>", update_comboboxes)
    Module_combobox.set('')  # 设置默认值

    # 创建一个动态框架来放置动态生成的部件
    dynamic_frame = ttk.Frame(root, padding="10 10 10 10")
    dynamic_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    return root

#配置OSS
def initialize_oss_client(config_file_path):
    # 检查配置文件是否存在
    if not os.path.exists(config_file_path):
        messagebox.showerror("提示", "系统检测缺失配置文件，请重启程序!\n\n\n若无法解决，请联系育种后台管理人员。")
        return None
    try:
        # 从配置文件中获取 OSS 配置信息
        access_key_id = config_dict['pass_word']['accesskeyid']
        access_key_secret = config_dict['pass_word']['accesskeysecret']
        bucket_name = config_dict['pass_word']['bucket_name']
        endpoint = config_dict['pass_word']['endpoint']

        # 初始化 OSS 客户端
        auth = oss2.Auth(access_key_id, access_key_secret)
        bucket = oss2.Bucket(auth, endpoint, bucket_name)

        # 返回 OSS Bucket 对象
        return bucket
    except Exception as e:
        messagebox.showerror("错误", f"发生了一个错误: {e}")

if getattr(sys, 'frozen', False):
    # 获取 PyInstaller 提取的临时文件路径
    base_path = sys._MEIPASS
else:
    # 否则使用脚本所在的路径
    base_path = os.path.abspath(".")
CONFIG_FILE_PATH = os.path.join(base_path, 'config.ini')
if __name__ == "__main__":
    #调用配置文件
    config_dict= config_load(CONFIG_FILE_PATH)
    if config_dict:
        # 调用OSS客户端
        bucket = initialize_oss_client(CONFIG_FILE_PATH)
        # 运行主窗口
        root = create_tk()
        root.mainloop()
    else:
        messagebox.showerror("提示", "系统检测缺失配置文件，请重启程序!\n\n\n若无法解决，请联系育种后台管理人员。")
